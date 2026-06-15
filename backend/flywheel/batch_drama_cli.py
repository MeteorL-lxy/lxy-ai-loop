from __future__ import annotations

import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import hashlib
import json
import os
from pathlib import Path
import random
import re
import subprocess
from types import SimpleNamespace
import sys
import time

from flywheel.clipping.ai_cut_animation import (
    DEFAULT_AUTO_CLIP_ENABLED,
    DEFAULT_MAX_EPISODES_PER_SERIAL,
    DEFAULT_MAX_TOTAL_DURATION_SECONDS,
    DEFAULT_PROCESS_CONCURRENCY,
    DEFAULT_SEGMENT_MAX_SECONDS,
    DEFAULT_SEGMENT_SECONDS,
    DEFAULT_SOURCE as DEFAULT_AI_ANIMATION_SOURCE,
    DEFAULT_SYNC_MATERIAL_AGENT_ID,
    DEFAULT_SYNC_MATERIAL_CATEGORY_ID,
    DEFAULT_TEMPLATE_ID,
    DEFAULT_USE_AUTO_MIGRATION,
)
from flywheel.selection.realtime_rank_source import mark_realtime_hour_exhausted


MODULE_ROOT_DIR = Path(__file__).resolve().parents[2]


def bind(ctx):
    protected = set(globals().keys())
    for name, value in vars(ctx).items():
        if name.startswith("__"):
            continue
        if name in protected and callable(globals().get(name)):
            continue
        globals()[name] = value
    return sys.modules[__name__]


def _resolve_batch_publish_targets(args: argparse.Namespace) -> dict[str, object]:
    platform = normalize_publish_platform(args.publish_platform)
    target = None
    account_pool_name = str(getattr(args, "account_pool", "") or "").strip()
    if not getattr(args, "account_id", None) and not getattr(args, "team_id", None) and not account_pool_name:
        if platform == "FACEBOOK":
            account_pool_name = DEFAULT_SHORT_DRAMA_FACEBOOK_ACCOUNT_POOL
    if not getattr(args, "account_id", None) and not getattr(args, "team_id", None) and not account_pool_name:
        accounts = require_success(get_publish_accounts(), "获取发布账号列表")
        matched_accounts = [
            account
            for account in accounts
            if str(account.get("type") or "").upper() == platform
            and str(account.get("team_id") or "").strip()
            and str(account.get("status")) == "0"
        ]
        if len(matched_accounts) == int(args.count):
            target = {
                "social_type": platform,
                "team_ids": [account.get("team_id") for account in matched_accounts],
                "accounts": matched_accounts,
            }

    if target is None:
        target = resolve_publish_targets(
            SimpleNamespace(
                account_id=args.account_id,
                team_id=args.team_id,
                account_pool=account_pool_name,
                platform=platform,
            )
        )
    social_type = str(target.get("social_type") or "").upper()
    if social_type != platform:
        raise SystemExit(f"账号平台与目标平台不一致: target={platform}, accounts={social_type}")

    accounts = target.get("accounts") if isinstance(target.get("accounts"), list) else []
    if accounts:
        items = [
            {
                "account_id": str(account.get("id") or ""),
                "team_id": str(account.get("team_id") or ""),
                "platform": social_type,
                "name": str(account.get("social_name") or "").strip() or f"{social_type} 账号",
            }
            for account in accounts
            if str(account.get("team_id") or "").strip()
        ]
    else:
        items = [
            {
                "account_id": "",
                "team_id": str(team_id),
                "platform": social_type,
                "name": f"{_platform_label(social_type)} 账号",
            }
            for team_id in target.get("team_ids", [])
            if str(team_id).strip()
        ]

    if not items:
        raise SystemExit("没有可用发布账号，请传 --account-id 或 --team-id")
    if len(items) < args.count and not args.allow_account_reuse:
        raise SystemExit(
            f"账号数量不足: 需要 {args.count} 个，当前 {len(items)} 个。"
            "如确认允许复用账号，请传 --allow-account-reuse。"
        )
    return {"platform": social_type, "accounts": items}


def _select_batch_dramas(args: argparse.Namespace, config, *, target_count: int | None = None) -> list[dict]:
    db = FlywheelSQLite(Path(config.database_path))
    db.init_schema(schema_path())
    desired_count = max(args.count, int(target_count or args.count))
    config.raw["flywheel"]["candidate_pool_size"] = max(
        int(config.raw["flywheel"].get("candidate_pool_size") or 30),
        max(desired_count * 15, 80),
    )
    language_mode = "all" if not args.language else "single"
    languages = [] if language_mode == "all" else [str(args.language)]
    requested_platform = str(args.drama_platform or "").strip().lower()
    if requested_platform:
        target_platforms = [CLIP_SUPPORTED_DRAMA_PLATFORM_ALIASES.get(requested_platform, requested_platform)]
        if target_platforms[0] not in CLIP_SUPPORTED_DRAMA_PLATFORMS:
            supported = "、".join(_app_label(item) for item in CLIP_SUPPORTED_DRAMA_PLATFORMS)
            raise SystemExit(f"{_app_label(requested_platform)} 暂不支持按剧集剪辑。当前支持剪辑的剧场: {supported}")
        platform_plan = [target_platforms[0]] * desired_count
        preferred_platforms = list(target_platforms)
    else:
        platform_plan = _random_batch_platform_plan(desired_count, CLIP_SUPPORTED_DRAMA_PLATFORMS)
        preferred_platforms = _unique_in_order(platform_plan)
        fallback_platforms = [item for item in CLIP_SUPPORTED_DRAMA_PLATFORMS if item not in preferred_platforms]
        random.shuffle(fallback_platforms)
        target_platforms = [*preferred_platforms, *fallback_platforms]

    candidates: list[dict] = []
    platform_counts = _count_platform_plan(platform_plan)
    original_pool_size = int(config.raw["flywheel"].get("candidate_pool_size") or max(desired_count * 4, 20))
    min_candidate_buffer = max(desired_count * 3, 12)
    fetched_preferred: set[str] = set()
    for platform in target_platforms:
        planned_slots = max(1, int(platform_counts.get(platform, 0) or 0))
        per_platform_pool_size = max(planned_slots * 12, 30)
        config.raw["flywheel"]["candidate_pool_size"] = per_platform_pool_size
        platform_candidates = fetch_candidates(
            config,
            platform=platform,
            languages=languages,
            language_mode_override=language_mode,
            order=args.drama_order,
            search=args.search or "",
        )
        for candidate in platform_candidates:
            row = dict(candidate)
            row["candidate_source_platform"] = platform
            candidates.append(row)
        eligible_count = sum(1 for item in candidates if _candidate_is_batch_eligible(item))
        if platform in preferred_platforms:
            fetched_preferred.add(platform)
        if eligible_count >= min_candidate_buffer and fetched_preferred >= set(preferred_platforms):
            break
    config.raw["flywheel"]["candidate_pool_size"] = original_pool_size
    eligible = [
        dict(item)
        for item in candidates
        if _candidate_is_batch_eligible(item)
    ]
    cooled_serial_ids = db.recent_learning_serial_ids(
        event_types=STRATEGY_MEMORY_EVENT_TYPES,
        recent_days=STRATEGY_MEMORY_COOLDOWN_DAYS,
    )
    cooled_eligible = [
        dict(item)
        for item in eligible
        if str(item.get("serial_id") or "").strip() not in cooled_serial_ids
    ]
    if len(cooled_eligible) >= desired_count:
        eligible = cooled_eligible
    recent_history_keys = db.recent_drama_pick_history_keys(
        recent_rounds=config.recent_hard_exclusion_rounds,
        recent_days=config.recent_hard_exclusion_days,
    )
    fresh_eligible, _ = split_recent_candidates(eligible, recent_history_keys)
    primary_pool = fresh_eligible if fresh_eligible else eligible
    primary_pool = _score_batch_candidates(primary_pool, db=db, config=config)
    primary_by_platform = _group_candidates_by_source_platform(primary_pool)
    selected: list[dict] = []
    used_identities: set[tuple[str, str, str]] = set()
    used_batch_keys: set[str] = set()

    for platform in platform_plan:
        rows = primary_by_platform.get(platform) or []
        while rows:
            candidate = _pick_light_random_candidate(rows)
            if candidate is None:
                break
            rows.remove(candidate)
            identity = _candidate_identity(candidate)
            batch_keys = _candidate_batch_keys(candidate)
            if identity in used_identities:
                continue
            if batch_keys and batch_keys & used_batch_keys:
                continue
            used_identities.add(identity)
            used_batch_keys.update(batch_keys)
            selected.append(candidate)
            break

    fallback_pool = [
        item
        for item in primary_pool
        if _candidate_identity(item) not in used_identities
        and not (_candidate_batch_keys(item) & used_batch_keys)
    ]
    fallback_pool = _score_batch_candidates(fallback_pool, db=db, config=config)
    _append_balanced_candidates(
        selected,
        fallback_pool,
        target_count=desired_count,
        platform_targets=platform_counts,
        used_identities=used_identities,
        used_batch_keys=used_batch_keys,
    )
    if len(selected) < desired_count and not fresh_eligible:
        extra_pool = [
            item
            for item in eligible
            if _candidate_identity(item) not in used_identities
            and not (_candidate_batch_keys(item) & used_batch_keys)
        ]
        extra_pool = _score_batch_candidates(extra_pool, db=db, config=config)
        _append_balanced_candidates(
            selected,
            extra_pool,
            target_count=desired_count,
            platform_targets=platform_counts,
            used_identities=used_identities,
            used_batch_keys=used_batch_keys,
        )

    if len(selected) < desired_count:
        if fresh_eligible:
            raise SystemExit(
                f"排除最近 {config.recent_hard_exclusion_rounds} 轮 / {config.recent_hard_exclusion_days} 天已选短剧后，可用新剧不足: "
                f"需要 {desired_count} 部，当前 {len(selected)} 部。"
            )
        raise SystemExit(f"可用短剧不足: 需要 {desired_count} 部，当前 {len(selected)} 部")
    return selected[:desired_count]


def _batch_episode_order(row: dict) -> int:
    return int(row.get("episode_order") or row.get("episode_id") or row.get("sequence") or row.get("id") or 0)


def _has_playable_episode_asset(row: dict, info=None) -> bool:
    info = info or {}
    return bool(row.get("play_url") or info.get("play_url") or info.get("mp4_OD") or info.get("m3u8_HD"))


def _select_or_validate_batch_episode(drama: dict, args: argparse.Namespace) -> dict:
    if _is_external_video_candidate(drama):
        external_video_url = _external_video_url(drama)
        duration_seconds = 0
        for value in (
            drama.get("external_video_duration_seconds"),
            ((drama.get("raw") or {}).get("external_video_duration_seconds") if isinstance(drama.get("raw"), dict) else 0),
        ):
            try:
                duration_seconds = max(duration_seconds, int(value or 0))
            except (TypeError, ValueError):
                continue
        return {
            "episode_order": 1,
            "episode_count": 1,
            "episode_id": 1,
            "episode_name": "Realtime External Video",
            "play_url": external_video_url,
            "duration": duration_seconds,
            "selection_mode": "external_video",
            "reason": "realtime_rank_external_video" if external_video_url else "missing_external_video_url",
            "supported": bool(external_video_url),
            "candidates": [],
        }
    serial_id = drama.get("serial_id")
    app_id = str(drama.get("app_id") or "")
    retry_count = max(0, int(getattr(args, "source_prepare_retry_count", 0) or 0))
    if not args.episode_order:
        try:
            return select_best_episode(serial_id, app_id, retry_count=retry_count)
        except InbeidouError as exc:
            return {
                "episode_order": 1,
                "episode_count": 0,
                "selection_mode": "unsupported_source",
                "reason": str(exc),
                "supported": False,
                "candidates": [],
            }

    forced_order = int(args.episode_order)
    try:
        rows = _episode_api_with_retries(
            lambda: require_success(get_episode_list(serial_id=int(serial_id)), "获取短剧剧集列表"),
            retry_count=retry_count,
        )
    except InbeidouError as exc:
        return {
            "episode_order": forced_order,
            "episode_count": 0,
            "selection_mode": "forced_episode_order",
            "reason": str(exc),
            "supported": False,
            "candidates": [],
        }

    episodes = [dict(row) for row in rows if _batch_episode_order(dict(row)) > 0]
    match = next((row for row in episodes if _batch_episode_order(row) == forced_order), {})
    if not match:
        return {
            "episode_order": forced_order,
            "episode_count": len(episodes),
            "selection_mode": "forced_episode_order",
            "reason": f"第 {forced_order} 集不存在或未返回剧集行",
            "supported": False,
            "candidates": [],
        }

    info: dict = {}
    if not match.get("play_url"):
        try:
            info = _episode_api_with_retries(
                lambda: require_success(
                    get_episode_info(
                        serial_id=int(serial_id),
                        episode_order=forced_order,
                        app_id=app_id,
                        task_type=str(drama.get("task_type") or "1"),
                    ),
                    f"获取第 {forced_order} 集详情",
                ),
                retry_count=retry_count,
            )
        except InbeidouError:
            info = {}
    if not _has_playable_episode_asset(match, info):
        return {
            "episode_order": forced_order,
            "episode_count": len(episodes),
            "selection_mode": "forced_episode_order",
            "reason": f"第 {forced_order} 集没有可剪辑的 mp4/play_url 素材",
            "supported": False,
            "candidates": [],
        }

    return {
        "episode_order": forced_order,
        "episode_count": len(episodes),
        "episode_id": match.get("episode_id") or info.get("id") or forced_order,
        "episode_name": match.get("episode_name") or info.get("chapter_name") or f"Episode {forced_order}",
        "play_url": match.get("play_url") or info.get("play_url") or info.get("mp4_OD") or info.get("m3u8_HD") or "",
        "duration": int(match.get("duration") or info.get("duration") or info.get("file_duration") or 0),
        "selection_mode": "forced_episode_order",
        "reason": "Forced episode was prechecked and has playable material for clipping.",
        "supported": True,
        "candidates": [],
    }


def _batch_episode_precheck_targets(dramas: list[dict], count: int, args: argparse.Namespace) -> list[dict]:
    del args
    target_window = min(
        BATCH_EPISODE_PRECHECK_MAX_WINDOW,
        max(BATCH_EPISODE_PRECHECK_MIN_WINDOW, int(count) + BATCH_EPISODE_PRECHECK_WINDOW_PADDING),
    )
    limit = min(len(dramas), target_window)
    return dramas[:limit]


def _select_batch_playable_dramas(
    args: argparse.Namespace,
    config,
    *,
    target_count: int | None = None,
) -> tuple[list[dict], list[dict[str, str]]]:
    selected: list[dict] = []
    skipped: list[dict[str, str]] = []
    desired_count = max(args.count, int(target_count or args.count))
    candidates = _select_batch_dramas(args, config, target_count=desired_count)
    cursor = 0

    while len(selected) < desired_count and cursor < len(candidates):
        window = _batch_episode_precheck_targets(candidates[cursor:], desired_count - len(selected), args)
        cursor += len(window)
        if not window:
            break
        max_workers = min(BATCH_EPISODE_PRECHECK_CONCURRENCY, max(1, len(window)))
        episode_results: dict[int, dict] = {}
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_map = {
                executor.submit(_select_or_validate_batch_episode, drama, args): (index, drama)
                for index, drama in enumerate(window)
            }
            for future in as_completed(future_map):
                index, drama = future_map[future]
                try:
                    episode = future.result()
                except Exception as exc:
                    episode = {
                        "supported": False,
                        "reason": str(exc),
                        "selection_mode": "episode_precheck_exception",
                    }
                episode_results[index] = {"drama": drama, "episode": episode}

        for index, drama in enumerate(window):
            result = episode_results.get(index) or {"drama": drama, "episode": {"supported": False, "reason": "missing_precheck_result"}}
            episode = result.get("episode") if isinstance(result.get("episode"), dict) else {}
            if not episode.get("supported"):
                skipped.append(
                    {
                        "serial_id": str(drama.get("serial_id") or ""),
                        "title": str(drama.get("title") or drama.get("title_ch") or ""),
                        "app_id": str(drama.get("app_id") or ""),
                        "reason": str(episode.get("reason") or episode.get("selection_mode") or "unsupported"),
                    }
                )
                continue
            selected.append(
                {
                    "drama": drama,
                    "episode": episode,
                }
            )
            if len(selected) >= desired_count:
                break

    return selected[:desired_count], skipped


def _strategy_memory_meta(config) -> dict[str, object]:
    db = FlywheelSQLite(Path(config.database_path))
    db.init_schema(schema_path())
    cooled_serial_ids = db.recent_learning_serial_ids(
        event_types=STRATEGY_MEMORY_EVENT_TYPES,
        recent_days=STRATEGY_MEMORY_COOLDOWN_DAYS,
    )
    return {
        "cooldown_days": STRATEGY_MEMORY_COOLDOWN_DAYS,
        "cooldown_serial_count": len(cooled_serial_ids),
        "event_types": list(STRATEGY_MEMORY_EVENT_TYPES),
    }


def _account_profile_lookup_keys(account: dict[str, object]) -> list[str]:
    keys: list[str] = []
    for value in (account.get("account_id"), account.get("name")):
        normalized = str(value or "").strip()
        if normalized and normalized not in keys:
            keys.append(normalized)
    return keys


def _lookup_account_assignment_profile(
    account: dict[str, object],
    profiles: dict[str, dict[str, object]],
) -> dict[str, object]:
    for key in _account_profile_lookup_keys(account):
        profile = profiles.get(key)
        if isinstance(profile, dict):
            return profile
    return {}


def _candidate_language_matches_profile(candidate: dict[str, object], profile: dict[str, object]) -> bool:
    preferred = profile.get("preferred_languages")
    if not isinstance(preferred, list) or not preferred:
        return True
    candidate_language = str(candidate.get("language") or "").strip()
    if not candidate_language:
        return False
    return candidate_language in {str(value).strip() for value in preferred if str(value).strip()}


def _candidate_blocked_by_profile(candidate: dict[str, object], profile: dict[str, object]) -> bool:
    blocked = profile.get("blocked_languages")
    if not isinstance(blocked, list) or not blocked:
        return False
    candidate_language = str(candidate.get("language") or "").strip()
    if not candidate_language:
        return False
    return candidate_language in {str(value).strip() for value in blocked if str(value).strip()}


def _candidate_hits_account_cooldown(candidate: dict[str, object], profile: dict[str, object]) -> bool:
    serial_ids = {str(value).strip() for value in (profile.get("recent_serial_ids") or []) if str(value).strip()}
    if serial_ids & candidate_serial_ids(candidate):
        return True
    title_keys = {str(value).strip() for value in (profile.get("recent_title_keys") or []) if str(value).strip()}
    return bool(title_keys & candidate_history_keys(candidate))


def _candidate_assignment_sort_key(
    candidate: dict[str, object],
    profile: dict[str, object],
) -> tuple[int, int, float, str]:
    cooldown_hit = _candidate_hits_account_cooldown(candidate, profile)
    language_match = _candidate_language_matches_profile(candidate, profile)
    blocked = _candidate_blocked_by_profile(candidate, profile)
    return (
        1 if cooldown_hit else 0,
        0 if language_match and not blocked else 1,
        -float(candidate.get("candidate_final_score") or 0.0),
        str(candidate.get("serial_id") or ""),
    )


def _assign_candidates_to_accounts(
    selected_sources: list[dict],
    accounts: list[dict],
    *,
    recent_days: int = 14,
) -> list[tuple[dict, dict]]:
    remaining = [dict(item) for item in selected_sources]
    profiles = build_account_assignment_profiles(recent_days=recent_days)
    assignments: list[tuple[dict, dict]] = []

    for account in accounts:
        if not remaining:
            break
        profile = _lookup_account_assignment_profile(account, profiles)
        chosen_index: int | None = None

        for require_fresh in (True, False):
            for require_language in (True, False):
                for index, candidate in enumerate(remaining):
                    if require_fresh and _candidate_hits_account_cooldown(candidate, profile):
                        continue
                    if require_language and not _candidate_language_matches_profile(candidate, profile):
                        continue
                    if _candidate_blocked_by_profile(candidate, profile):
                        continue
                    chosen_index = index
                    break
                if chosen_index is not None:
                    break
            if chosen_index is not None:
                break

        if chosen_index is None:
            remaining.sort(key=lambda item: _candidate_assignment_sort_key(item, profile))
            chosen = remaining.pop(0)
        else:
            chosen = remaining.pop(chosen_index)

        assignments.append((account, chosen))

    return assignments


def _clone_reused_batch_source(source: dict, *, reuse_index: int) -> dict[str, object]:
    cloned = {
        "drama": dict(source.get("drama") or {}),
        "episode": dict(source.get("episode") or {}),
        "source_reused": True,
        "reuse_index": max(1, int(reuse_index or 1)),
    }
    if str(((cloned.get("drama") or {}).get("source_mode") or "")).strip().lower() != "external_video":
        cloned["episode_rotation_index"] = max(1, int(reuse_index or 1)) + 1
    return cloned


def _source_identity(source: dict) -> tuple[str, str, str]:
    drama = source.get("drama") if isinstance(source.get("drama"), dict) else {}
    return _candidate_identity(drama)


def _source_is_external_video(source: dict) -> bool:
    drama = source.get("drama") if isinstance(source.get("drama"), dict) else {}
    source_mode = str(drama.get("source_mode") or "").strip().lower()
    return source_mode in {"external_video", "official_ffmpeg"} or _is_external_video_candidate(drama)


def _source_external_output_capacity(source: dict) -> int:
    drama = source.get("drama") if isinstance(source.get("drama"), dict) else {}
    episode = source.get("episode") if isinstance(source.get("episode"), dict) else {}
    duration_seconds = 0
    for value in (
        drama.get("external_video_duration_seconds"),
        ((drama.get("raw") or {}).get("external_video_duration_seconds") if isinstance(drama.get("raw"), dict) else 0),
        episode.get("duration"),
    ):
        try:
            duration_seconds = max(duration_seconds, int(value or 0))
        except (TypeError, ValueError):
            continue
    if duration_seconds > 0:
        return max(1, min(60, duration_seconds // 15))
    for value in (
        drama.get("external_estimated_output_count"),
        ((drama.get("raw") or {}).get("external_estimated_output_count") if isinstance(drama.get("raw"), dict) else 0),
    ):
        try:
            count = int(value or 0)
        except (TypeError, ValueError):
            continue
        if count > 0:
            return max(1, min(60, count))
    return 1


def _prioritize_batch_sources(
    sources: list[dict],
    *,
    requested_count: int,
) -> tuple[list[dict], list[dict], dict[str, int]]:
    requested = max(1, int(requested_count or 0))
    primary: list[dict] = []
    used_unique_keys: set[tuple[str, str, str]] = set()
    realtime_external_unique_count = 0
    realtime_external_slot_fill_count = 0

    for source in sources:
        if not _source_is_external_video(source):
            continue
        realtime_external_unique_count += 1
        copies = _source_external_output_capacity(source)
        for copy_index in range(copies):
            if len(primary) >= requested:
                break
            if copy_index == 0:
                primary.append(dict(source))
                used_unique_keys.add(_source_identity(source))
            else:
                primary.append(_clone_reused_batch_source(source, reuse_index=copy_index))
            realtime_external_slot_fill_count += 1
        if len(primary) >= requested:
            break

    for source in sources:
        if len(primary) >= requested:
            break
        identity = _source_identity(source)
        if identity in used_unique_keys:
            continue
        primary.append(dict(source))
        used_unique_keys.add(identity)

    reserve = [
        dict(source)
        for source in sources
        if _source_identity(source) not in used_unique_keys
    ]
    meta = {
        "realtime_external_unique_count": realtime_external_unique_count,
        "realtime_external_slot_fill_count": realtime_external_slot_fill_count,
    }
    return primary[:requested], reserve, meta


def _expand_batch_sources_with_reuse(
    sources: list[dict],
    *,
    requested_count: int,
) -> tuple[list[dict], int]:
    if len(sources) >= requested_count or not sources:
        return [dict(source) for source in sources[:requested_count]], 0
    expanded = [dict(source) for source in sources]
    reuse_count = 0
    while len(expanded) < requested_count:
        template = sources[reuse_count % len(sources)]
        expanded.append(_clone_reused_batch_source(template, reuse_index=reuse_count + 1))
        reuse_count += 1
    return expanded[:requested_count], reuse_count


def _normalize_external_clip_duration(source: dict, requested_duration) -> int | str:
    line_name = _line_name()
    raw_requested = str(requested_duration or "").strip().lower()
    if raw_requested and raw_requested != "auto":
        try:
            explicit = int(float(raw_requested))
        except (TypeError, ValueError):
            explicit = 0
        if explicit > 0:
            if line_name in {"realtime", "realtime_day", "realtime_single"}:
                return max(20, min(30, explicit))
            if line_name == "yourchannel":
                return max(15, min(20, explicit))
            return max(12, min(20, explicit))
    if line_name in {"realtime", "realtime_day", "realtime_single"}:
        return 25
    if line_name == "yourchannel":
        return 18
    return 15


def _normalize_official_ffmpeg_clip_duration(requested_duration) -> int | str:
    line_name = _line_name()
    raw_requested = str(requested_duration or "").strip().lower()
    if raw_requested and raw_requested != "auto":
        try:
            explicit = int(float(raw_requested))
        except (TypeError, ValueError):
            explicit = 0
        if explicit > 0:
            if line_name == "yourchannel":
                return max(15, min(20, explicit))
            return max(12, min(20, explicit))
    if line_name == "yourchannel":
        return 18
    return 15


def _resolve_batch_clip_duration(source: dict, requested_duration) -> int | str:
    drama = source.get("drama") if isinstance(source.get("drama"), dict) else {}
    source_mode = str(drama.get("source_mode") or "").strip().lower()
    if source_mode == "external_video":
        return _normalize_external_clip_duration(source, requested_duration)
    if source_mode == "official_ffmpeg":
        return _normalize_official_ffmpeg_clip_duration(requested_duration)
    return requested_duration


def _line_name() -> str:
    return str(os.getenv("BARRY_LOOP_LINE_NAME") or "").strip().lower()


def _line_in(*names: str) -> bool:
    return _line_name() in {str(item).strip().lower() for item in names if str(item).strip()}


def _realtime_material_mode_enabled() -> bool:
    line_name = _line_name()
    if line_name not in {"realtime", "realtime_day", "realtime_single"}:
        return False
    value = str(os.getenv("BARRY_LOOP_REALTIME_MATERIAL_ONLY") or "0").strip().lower()
    return value in {"1", "true", "yes", "on"}


def _creative_list_material_mode_enabled() -> bool:
    if not _line_in("creative_list", "creative_list_day"):
        return False
    value = str(os.getenv("BARRY_LOOP_CREATIVE_LIST_MATERIAL_ONLY") or "0").strip().lower()
    return value in {"1", "true", "yes", "on"}


def _official_ffmpeg_mode_enabled() -> bool:
    return _line_in("fbhot_test", "yourchannel")


def _yourchannel_mode_enabled() -> bool:
    return _line_in("yourchannel")


def _select_realtime_external_sources(args: argparse.Namespace, config, *, target_count: int | None = None) -> tuple[list[dict], list[dict]]:
    desired_count = max(args.count, int(target_count or args.count))
    publish_platform = normalize_publish_platform(args.publish_platform)
    fetch_target_size = max(desired_count * 6, 30)
    requested_platform = str(args.drama_platform or "").strip().lower()
    external_candidates = []
    skipped: list[dict[str, object]] = []

    for row in fetch_realtime_rank_candidates(
        config,
        target_publish_platforms=[publish_platform],
        target_size=fetch_target_size,
    ):
        candidate = dict(row)
        app_id = str(candidate.get("app_id") or "").strip().lower()
        if requested_platform and app_id != requested_platform:
            skipped.append(
                {
                    "title": str(candidate.get("title") or ""),
                    "app_id": app_id,
                    "candidate_fetch_source": str(candidate.get("candidate_fetch_source") or ""),
                    "reason": "drama_platform_filtered",
                }
            )
            continue
        if not _is_external_video_candidate(candidate):
            skipped.append(
                {
                    "title": str(candidate.get("title") or ""),
                    "app_id": app_id,
                    "candidate_fetch_source": str(candidate.get("candidate_fetch_source") or ""),
                    "reason": "non_external_realtime_candidate",
                }
            )
            continue
        if not _external_video_url(candidate):
            skipped.append(
                {
                    "title": str(candidate.get("title") or ""),
                    "app_id": app_id,
                    "candidate_fetch_source": str(candidate.get("candidate_fetch_source") or ""),
                    "reason": "missing_external_video_url",
                }
            )
            continue
        external_candidates.append(candidate)

    if not external_candidates:
        mark_realtime_hour_exhausted(
            "当前小时缓存素材已耗尽或未命中可下载外部素材",
            target_publish_platforms=[publish_platform],
        )
        raise SystemExit(
            "实时榜线路当前没有可下载外部素材；已跳过普通短剧候选池，不回退到官方选剧逻辑。"
        )

    sources: list[dict[str, object]] = []
    preview_skipped = skipped[:10]
    for drama in external_candidates:
        episode = _select_or_validate_batch_episode(drama, args)
        if not bool(episode.get("supported")):
            preview_skipped.append(
                {
                    "title": str(drama.get("title") or ""),
                    "app_id": str(drama.get("app_id") or ""),
                    "candidate_fetch_source": str(drama.get("candidate_fetch_source") or ""),
                    "reason": str(episode.get("reason") or "unsupported_external_video"),
                }
            )
            continue
        sources.append({"drama": drama, "episode": episode})

    if not sources:
        mark_realtime_hour_exhausted(
            "当前小时缓存候选均无法直接进入剪辑",
            target_publish_platforms=[publish_platform],
        )
        raise SystemExit(
            "实时榜线路已拉到外部素材候选，但没有可直接进入剪辑的素材；本轮不回退到官方选剧逻辑。"
        )
    return sources, preview_skipped


def _select_creative_list_external_sources(args: argparse.Namespace, config, *, target_count: int | None = None) -> tuple[list[dict], list[dict]]:
    desired_count = max(args.count, int(target_count or args.count))
    publish_platform = normalize_publish_platform(args.publish_platform)
    fetch_target_size = max(desired_count * 4, 40)
    candidates = fetch_creative_list_candidates(
        config,
        target_publish_platforms=[publish_platform],
        target_size=fetch_target_size,
        line_name=_line_name(),
    )
    if not candidates:
        raise SystemExit(
            "创意列表线路当前没有匹配到可下载外部素材；已完成本轮剧场扫描，不回退到官方选剧逻辑。"
        )

    sources: list[dict[str, object]] = []
    skipped: list[dict[str, object]] = []
    for drama in candidates:
        episode = _select_or_validate_batch_episode(drama, args)
        if not bool(episode.get("supported")):
            skipped.append(
                {
                    "title": str(drama.get("title") or ""),
                    "app_id": str(drama.get("app_id") or ""),
                    "candidate_fetch_source": str(drama.get("candidate_fetch_source") or ""),
                    "reason": str(episode.get("reason") or "unsupported_external_video"),
                }
            )
            continue
        sources.append({"drama": drama, "episode": episode})
    if not sources:
        raise SystemExit(
            "创意列表线路已匹配到候选素材，但没有可直接进入剪辑的外部视频。"
        )
    return sources, skipped[:10]


def _yourchannel_title_allowlist_path() -> Path:
    raw = str(os.getenv("BARRY_LOOP_YOURCHANNEL_ALLOWLIST_FILE") or "").strip()
    if raw:
        return Path(raw).expanduser().resolve()
    preferred_excel = Path(
        "/Users/xinyuliu/Downloads/创作者上传英文剧看板_交叉表_20260610.xlsx"
    )
    if preferred_excel.exists():
        return preferred_excel.resolve()
    return (MODULE_ROOT_DIR / "data" / "yourchannel_titles.json").resolve()


def _load_yourchannel_titles_from_excel(path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    worksheet = workbook[workbook.sheetnames[0]]
    title_index = 3
    try:
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        for idx, value in enumerate(header):
            if str(value or "").strip() == "剧名":
                title_index = idx
                break
    except Exception:
        pass
    titles: list[str] = []
    seen: set[str] = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if title_index >= len(row):
            continue
        title = str(row[title_index] or "").strip()
        if not title or title in seen:
            continue
        seen.add(title)
        titles.append(title)
    return titles


def _load_yourchannel_titles() -> list[str]:
    path = _yourchannel_title_allowlist_path()
    if not path.exists():
        return []
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _load_yourchannel_titles_from_excel(path)
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        payload = []
    if isinstance(payload, list):
        return [str(item).strip() for item in payload if str(item).strip()]
    return []


def _normalize_title_for_match(value: str) -> str:
    normalized = str(value or "").strip().lower()
    normalized = normalized.replace("’", "'").replace("‘", "'")
    normalized = normalized.replace("“", '"').replace("”", '"')
    normalized = re.sub(r"[^\w]+", " ", normalized, flags=re.UNICODE)
    return " ".join(normalized.split())


def _find_official_title_match(title: str, app_id: str) -> dict:
    body = require_success(
        get_tasks(
            page=1,
            page_size=20,
            platform=app_id,
            language="",
            search=title,
            order="publish_at",
            task_type="1",
        ),
        f"查找官方短剧 {title}",
    )
    rows = body.get("data", []) if isinstance(body, dict) else []
    wanted = _normalize_title_for_match(title)
    for row in rows:
        if not isinstance(row, dict):
            continue
        titles = [
            str(row.get("title") or "").strip(),
            str(row.get("title_ch") or "").strip(),
            str(row.get("title_en") or "").strip(),
        ]
        if any(_normalize_title_for_match(item) == wanted for item in titles if item):
            return dict(row)
    return {}


def _select_yourchannel_sources(args: argparse.Namespace, config, *, target_count: int | None = None) -> tuple[list[dict], list[dict]]:
    del config
    desired_count = max(args.count, int(target_count or args.count))
    titles = _load_yourchannel_titles()
    if not titles:
        raise SystemExit(
            f"YourChannel 白名单为空或缺失：{_yourchannel_title_allowlist_path()}"
        )
    sources: list[dict[str, object]] = []
    skipped: list[dict[str, object]] = []
    for title in titles:
        drama = _find_official_title_match(title, "yourchannel_drama")
        if not drama:
            skipped.append({"title": title, "app_id": "yourchannel_drama", "reason": "official_title_not_found"})
            continue
        episode = _select_or_validate_batch_episode(drama, args)
        if not bool(episode.get("supported")):
            skipped.append(
                {
                    "title": title,
                    "app_id": "yourchannel_drama",
                    "reason": str(episode.get("reason") or "unsupported_episode"),
                }
            )
            continue
        drama["source_mode"] = "official_ffmpeg"
        drama["candidate_fetch_source"] = "yourchannel_official_ffmpeg"
        sources.append({"drama": drama, "episode": episode})
        if len(sources) >= desired_count:
            break
    if not sources:
        raise SystemExit(
            "YourChannel 白名单剧名已扫描完成，但没有可直接进入 ffmpeg 剪辑发布的官方剧集。"
        )
    return sources, skipped[:10]


def _expand_ordinary_sources_with_episode_rotation(
    sources: list[dict],
    *,
    requested_count: int,
) -> list[dict]:
    if not sources:
        return []
    expanded: list[dict] = []
    episode_indices = [0 for _ in sources]
    episode_limits = [
        max(1, int(((source.get("episode") or {}).get("episode_count") or 1)))
        for source in sources
    ]

    while len(expanded) < requested_count:
        progressed = False
        for index, source in enumerate(sources):
            if len(expanded) >= requested_count:
                break
            episode_index = episode_indices[index]
            if episode_index >= episode_limits[index]:
                continue
            if episode_index == 0:
                item = dict(source)
                item["episode_rotation_index"] = 1
            else:
                item = _clone_reused_batch_source(source, reuse_index=episode_index)
                item["episode_rotation_index"] = episode_index + 1
            expanded.append(item)
            episode_indices[index] += 1
            progressed = True
        if not progressed:
            break
    return expanded[:requested_count]


def _resolve_ordinary_round_robin_episode(
    drama: dict,
    episode: dict,
    args: argparse.Namespace,
    *,
    desired_order: int,
) -> dict:
    desired = max(1, int(desired_order or 1))
    episode_count = max(1, int(episode.get("episode_count") or desired))
    cached: dict[int, dict] = {}

    def _pick(order: int) -> dict:
        order = max(1, int(order or 1))
        if order in cached:
            return dict(cached[order])
        forced_args = SimpleNamespace(**vars(args))
        forced_args.episode_order = order
        resolved = _select_or_validate_batch_episode(drama, forced_args)
        cached[order] = dict(resolved)
        return dict(resolved)

    for order in range(desired, episode_count + 1):
        resolved = _pick(order)
        if bool(resolved.get("supported")):
            resolved["selection_mode"] = "round_robin_episode_order"
            resolved["reason"] = f"Round-robin episode order assigned to account slot: E{order:02d}"
            return resolved
    for order in range(1, desired):
        resolved = _pick(order)
        if bool(resolved.get("supported")):
            resolved["selection_mode"] = "round_robin_episode_order_wrap"
            resolved["reason"] = f"Round-robin episode order wrapped to playable episode: E{order:02d}"
            return resolved
    fallback = dict(episode)
    fallback["selection_mode"] = "round_robin_episode_order_fallback"
    fallback["reason"] = f"Round-robin episode order fallback to prechecked episode after no playable forced order near E{desired:02d}"
    return fallback


def _build_batch_plan_item(
    *,
    slot_index: int,
    account: dict,
    source: dict,
    duration: int,
    args: argparse.Namespace | None = None,
) -> dict[str, object]:
    drama = source["drama"]
    episode = dict(source["episode"])
    source_mode = str(drama.get("source_mode") or "official")
    resolved_duration = _resolve_batch_clip_duration(source, duration)
    if source_mode in {"external_video", "official_ffmpeg"}:
        clip_options = {
            "provider": "external_video_ffmpeg_segment" if source_mode == "external_video" else "official_ffmpeg_segment",
            "cut_type": "ffmpeg_segment",
            "duration": resolved_duration,
            "output_count": 1,
            "script_count": 1,
            "deduplication": [],
            "watermark": False,
            "target_aspect_ratio": "9:16",
        }
    else:
        preferred_episode_order = max(0, int(episode.get("episode_order") or 0))
        clip_options = {
            "provider": "ai_cut_animation",
            "cut_type": "ai_cut_animation",
            "duration": f"{DEFAULT_SEGMENT_SECONDS}-{DEFAULT_SEGMENT_MAX_SECONDS}s",
            "output_count": 1,
            "script_count": 1,
            "deduplication": [],
            "watermark": False,
            "target_aspect_ratio": "9:16",
            "template_id": DEFAULT_TEMPLATE_ID,
            "segment_seconds": DEFAULT_SEGMENT_SECONDS,
            "segment_max_seconds": DEFAULT_SEGMENT_MAX_SECONDS,
            "process_concurrency": DEFAULT_PROCESS_CONCURRENCY,
            "max_total_duration_seconds": DEFAULT_MAX_TOTAL_DURATION_SECONDS,
            "max_episodes_per_serial": max(DEFAULT_MAX_EPISODES_PER_SERIAL, preferred_episode_order),
            "source": DEFAULT_AI_ANIMATION_SOURCE,
            "auto_clip_enabled": DEFAULT_AUTO_CLIP_ENABLED,
            "use_auto_migration": DEFAULT_USE_AUTO_MIGRATION,
            "auto_clip_sync_material_agent_id": DEFAULT_SYNC_MATERIAL_AGENT_ID,
            "auto_clip_sync_material_category_id": DEFAULT_SYNC_MATERIAL_CATEGORY_ID,
        }
        desired_episode_order = max(1, int(source.get("episode_rotation_index") or 0))
        if desired_episode_order > 0 and args is not None:
            episode = _resolve_ordinary_round_robin_episode(
                drama,
                episode,
                args,
                desired_order=desired_episode_order,
            )
    return {
        "index": slot_index + 1,
        "drama": {
            "serial_id": drama.get("serial_id"),
            "task_id": drama.get("task_id"),
            "task_type": drama.get("task_type") or "1",
            "title": drama.get("title") or drama.get("title_ch") or "",
            "app_id": drama.get("app_id"),
            "source_platform": drama.get("candidate_source_platform") or drama.get("app_id"),
            "language": str(drama.get("language") or ""),
            "share_rate": drama.get("share_rate"),
            "candidate_fetch_source": _candidate_fetch_source(drama) or "task_api",
            "candidate_final_score": float(drama.get("candidate_final_score") or 0.0),
            "candidate_score_breakdown": dict(drama.get("candidate_score_breakdown") or {}),
            "source_mode": source_mode,
            "external_video_url": _external_video_url(drama),
            "external_video_duration_seconds": int(drama.get("external_video_duration_seconds") or 0),
            "external_estimated_output_count": int(drama.get("external_estimated_output_count") or 0),
            "promotion_anchor": dict(drama.get("promotion_anchor") or {}) if isinstance(drama.get("promotion_anchor"), dict) else {},
            "matched_official_task_id": str(drama.get("matched_official_task_id") or ""),
            "matched_official_serial_id": str(drama.get("matched_official_serial_id") or ""),
            "matched_official_language": str(drama.get("matched_official_language") or ""),
            "source_reused": bool(source.get("source_reused")),
            "reuse_index": int(source.get("reuse_index") or 0),
            "episode_rotation_index": int(source.get("episode_rotation_index") or 0),
            "raw": drama,
        },
        "episode": episode,
        "account": account,
        "clip_options": clip_options,
    }


def _build_batch_replacement_item(slot_item: dict, source: dict, *, attempt_no: int) -> dict[str, object]:
    clip_options = slot_item.get("clip_options") if isinstance(slot_item.get("clip_options"), dict) else {}
    raw_duration = clip_options.get("duration")
    try:
        duration_value = int(raw_duration or 0)
    except Exception:
        duration_value = 0
    replacement = _build_batch_plan_item(
        slot_index=max(0, int(slot_item.get("index") or 1) - 1),
        account=dict(slot_item.get("account") or {}),
        source=source,
        duration=duration_value,
    )
    replacement["replacement_for_index"] = int(slot_item.get("index") or 0)
    replacement["replacement_attempt"] = max(1, int(attempt_no or 1))
    return replacement


def _build_batch_plan(args: argparse.Namespace, config) -> dict[str, object]:
    target = _resolve_batch_publish_targets(args)
    accounts = list(target["accounts"])
    random.shuffle(accounts)
    if args.allow_account_reuse and len(accounts) < args.count:
        expanded_accounts = []
        for index in range(args.count):
            expanded_accounts.append(accounts[index % len(accounts)])
        accounts = expanded_accounts
    else:
        accounts = accounts[: args.count]

    plan_items: list[dict] = []
    if _yourchannel_mode_enabled():
        selected_sources, skipped = _select_yourchannel_sources(
            args,
            config,
            target_count=_batch_source_reserve_target(args.count),
        )
    elif _creative_list_material_mode_enabled():
        selected_sources, skipped = _select_creative_list_external_sources(
            args,
            config,
            target_count=_batch_source_reserve_target(args.count),
        )
    elif _realtime_material_mode_enabled():
        selected_sources, skipped = _select_realtime_external_sources(
            args,
            config,
            target_count=_batch_source_reserve_target(args.count),
        )
    else:
        selected_sources, skipped = _select_batch_playable_dramas(
            args,
            config,
            target_count=_batch_source_reserve_target(args.count),
        )
        if _official_ffmpeg_mode_enabled():
            selected_sources = [
                {
                    **dict(source),
                    "drama": {
                        **dict(source.get("drama") or {}),
                        "source_mode": "official_ffmpeg",
                        "candidate_fetch_source": str(((source.get("drama") or {}).get("candidate_fetch_source")) or "official_ffmpeg"),
                    },
                }
                for source in selected_sources
            ]
    primary_sources, reserve_sources, source_priority_meta = _prioritize_batch_sources(
        selected_sources,
        requested_count=args.count,
    )
    unique_playable_source_count = len(
        {
            _source_identity(source)
            for source in primary_sources
            if _source_identity(source) != ("", "", "")
        }
    )
    if _realtime_material_mode_enabled():
        primary_sources, source_reuse_fill_count = _expand_batch_sources_with_reuse(
            primary_sources,
            requested_count=len(accounts),
        )
    else:
        primary_sources = _expand_ordinary_sources_with_episode_rotation(
            primary_sources,
            requested_count=len(accounts),
        )
        source_reuse_fill_count = max(0, len(primary_sources) - unique_playable_source_count)
    assignments = _assign_candidates_to_accounts(primary_sources, accounts)
    for account, source in assignments:
        index = len(plan_items)
        plan_items.append(
            _build_batch_plan_item(
                slot_index=index,
                account=account,
                source=source,
                duration=args.duration,
                args=args,
            )
        )
        if len(plan_items) >= args.count:
            break
    return {
        "platform": target["platform"],
        "items": plan_items,
        "drama_platform_plan": [
            str((item.get("drama") or {}).get("source_platform") or (item.get("drama") or {}).get("app_id") or "")
            for item in plan_items
        ],
        "episode_precheck": {
            "concurrency": BATCH_EPISODE_PRECHECK_CONCURRENCY,
            "window_padding": BATCH_EPISODE_PRECHECK_WINDOW_PADDING,
            "min_window": BATCH_EPISODE_PRECHECK_MIN_WINDOW,
            "max_window": BATCH_EPISODE_PRECHECK_MAX_WINDOW,
        },
        "strategy_memory": _strategy_memory_meta(config),
        "skipped_preview": skipped[:10],
        "reserve_sources": reserve_sources,
        "replacement_buffer": len(reserve_sources),
        "unique_playable_source_count": unique_playable_source_count,
        "source_reuse_fill_count": source_reuse_fill_count,
        "realtime_external_unique_count": int(source_priority_meta.get("realtime_external_unique_count") or 0),
        "realtime_external_slot_fill_count": int(source_priority_meta.get("realtime_external_slot_fill_count") or 0),
        "planned_shortfall_count": max(0, int(args.count) - len(plan_items)),
    }


def _batch_safety_gate_item(item: dict, platform: str, *, prefetch_promotion: bool) -> dict:
    account = item.get("account") or {}
    drama = item.get("drama") or {}
    episode = item.get("episode") or {}
    team_id = str(account.get("team_id") or "").strip()
    if not team_id:
        return {
            **item,
            "safety_gate": {"passed": False, "reason": "缺少可发布 team_id"},
        }
    if not str(episode.get("play_url") or "").strip():
        return {
            **item,
            "safety_gate": {"passed": False, "reason": "剧集缺少可剪辑 play_url/mp4 素材"},
        }
    try:
        validate_source_episode_constraints(
            episode,
            source_mode=str(drama.get("source_mode") or ""),
        )
    except Exception as exc:
        return {
            **item,
            "safety_gate": {"passed": False, "reason": f"源剧集约束校验失败: {exc}"},
        }
    promotion = item.get("promotion") if isinstance(item.get("promotion"), dict) else {}
    if prefetch_promotion:
        try:
            promotion = _promotion_caption(item, platform)
            validate_promotion_constraints(platform, promotion)
        except Exception as exc:
            return {
                **item,
                "safety_gate": {"passed": False, "reason": f"推广文案/链接获取失败: {exc}"},
            }
        if PUBLISH_TO_PROMOTION_PLATFORM.get(platform) and not str(promotion.get("promotion_link") or "").strip():
            title = str(drama.get("title") or "")
            return {
                **item,
                "safety_gate": {"passed": False, "reason": f"《{title or '短剧'}》未拿到可发布推广链接"},
            }
    return {
        **item,
        "promotion": promotion,
        "safety_gate": {
            "passed": True,
            "reason": "已通过安全门槛",
        },
    }


def _run_batch_safety_gate(items: list[dict], platform: str, *, prefetch_promotion: bool) -> tuple[list[dict], list[dict]]:
    checked = _run_parallel(
        items,
        min(BATCH_SAFETY_GATE_CONCURRENCY, max(1, len(items))),
        lambda item: _batch_safety_gate_item(item, platform, prefetch_promotion=prefetch_promotion),
    )
    approved = [item for item in checked if bool((item.get("safety_gate") or {}).get("passed"))]
    rejected = [item for item in checked if not bool((item.get("safety_gate") or {}).get("passed"))]
    approved.sort(key=lambda item: int(item.get("index") or 0))
    rejected.sort(key=lambda item: int(item.get("index") or 0))
    return approved, rejected


def _format_batch_safety_reject(item: dict, *, replaced: bool | None = None) -> dict[str, object]:
    drama = item.get("drama") if isinstance(item.get("drama"), dict) else {}
    source_platform = str(drama.get("source_platform") or drama.get("app_id") or "").strip()
    payload: dict[str, object] = {
        "短剧": str(drama.get("title") or "").strip(),
        "短剧ID": str(drama.get("serial_id") or "").strip(),
        "剧场": _app_label(source_platform),
        "集数": int(((item.get("episode") or {}).get("episode_order") or 0)),
        "原因": str((item.get("safety_gate") or {}).get("reason") or "").strip(),
    }
    if replaced is not None:
        payload["补位结果"] = "已补位" if replaced else "未补位"
    if int(item.get("replacement_for_index") or 0) > 0:
        payload["类型"] = "补位候选"
        payload["原槽位"] = int(item.get("replacement_for_index") or 0)
    else:
        payload["类型"] = "原计划候选"
        payload["原槽位"] = int(item.get("index") or 0)
    return payload


def _backfill_batch_safety_gate(
    *,
    approved_items: list[dict],
    rejected_items: list[dict],
    reserve_sources: list[dict],
    args: argparse.Namespace,
    platform: str,
    prefetch_promotion: bool,
) -> tuple[list[dict], list[dict], dict[str, object]]:
    final_items = list(approved_items)
    all_rejected = list(rejected_items)
    pending_slots = sorted(rejected_items, key=lambda item: int(item.get("index") or 0))
    reserve_queue = list(reserve_sources)
    reserve_attempts = 0
    replacements_filled = 0
    reject_results: list[dict[str, object]] = [_format_batch_safety_reject(item, replaced=False) for item in pending_slots]
    reject_result_by_slot = {
        int(item.get("原槽位") or 0): item
        for item in reject_results
        if int(item.get("原槽位") or 0) > 0
    }

    while pending_slots and reserve_queue and len(final_items) < int(args.count):
        slot_item = pending_slots.pop(0)
        source = reserve_queue.pop(0)
        reserve_attempts += 1
        replacement_item = _build_batch_replacement_item(slot_item, source, attempt_no=reserve_attempts)
        approved_batch, rejected_batch = _run_batch_safety_gate(
            [replacement_item],
            platform,
            prefetch_promotion=prefetch_promotion,
        )
        if approved_batch:
            final_items.extend(approved_batch)
            replacements_filled += 1
            slot_key = int(slot_item.get("index") or 0)
            if slot_key in reject_result_by_slot:
                reject_result_by_slot[slot_key]["补位结果"] = "已补位"
            continue
        if rejected_batch:
            rejected = rejected_batch[0]
            all_rejected.append(rejected)
            reject_results.append(_format_batch_safety_reject(rejected, replaced=False))
            pending_slots.insert(0, slot_item)

    final_items.sort(key=lambda item: int(item.get("index") or 0))
    unfilled_slots = [int(item.get("index") or 0) for item in pending_slots if int(item.get("index") or 0) > 0]
    for slot in unfilled_slots:
        if slot in reject_result_by_slot:
            reject_result_by_slot[slot]["补位结果"] = "未补位"
    meta = {
        "requested_count": int(args.count),
        "reserve_source_count": len(reserve_sources),
        "reserve_attempt_count": reserve_attempts,
        "replacement_filled_count": replacements_filled,
        "unfilled_count": max(0, int(args.count) - len(final_items)),
        "unfilled_slots": unfilled_slots,
        "rejected_details": reject_results,
    }
    return final_items[: int(args.count)], all_rejected, meta


def _record_batch_learning_logs(*, round_id: int, items: list[dict], safety_rejected: list[dict], report_zh: dict, config) -> None:
    db = FlywheelSQLite(Path(config.database_path))
    db.init_schema(schema_path())
    rows: list[dict] = []
    for item in safety_rejected:
        drama = item.get("drama") or {}
        rows.append(
            {
                "round_id": round_id,
                "event_type": "safety_reject",
                "serial_id": str(drama.get("serial_id") or ""),
                "payload": {
                    "title": str(drama.get("title") or ""),
                    "app_id": str(drama.get("app_id") or drama.get("source_platform") or ""),
                    "episode_order": int(((item.get("episode") or {}).get("episode_order") or 0)),
                    "reason": str((item.get("safety_gate") or {}).get("reason") or ""),
                },
            }
        )
    for item in items:
        drama = item.get("drama") or {}
        clip_options = item.get("clip_options") or {}
        event_type = "publish_processing"
        error_text = str(item.get("error") or "").strip()
        if "等待短剧素材就绪超时" in error_text:
            event_type = "clip_failed_source_prepare"
        elif item.get("status") == "failed":
            event_type = "publish_failed_final"
        elif str((item.get("publish") or {}).get("tasks") or "") and any(
            str(report.get("短剧ID") or "") == str(drama.get("serial_id") or "")
            and str(report.get("账号") or "") == str((item.get("account") or {}).get("name") or "")
            and str(report.get("发布情况") or "") == "发布成功"
            for report in (report_zh.get("任务明细") if isinstance(report_zh.get("任务明细"), list) else [])
        ):
            event_type = "publish_success"
        rows.append(
            {
                "round_id": round_id,
                "event_type": event_type,
                "serial_id": str(drama.get("serial_id") or ""),
                "payload": {
                    "title": str(drama.get("title") or ""),
                    "app_id": str(drama.get("app_id") or drama.get("source_platform") or ""),
                    "language": str(drama.get("language") or ""),
                    "episode_order": int(((item.get("episode") or {}).get("episode_order") or 0)),
                    "account": str((item.get("account") or {}).get("name") or ""),
                    "platform": str((item.get("account") or {}).get("platform") or report_zh.get("目标平台") or ""),
                    "cut_type": str(clip_options.get("cut_type") or ""),
                    "deduplication": list(clip_options.get("deduplication") or []),
                    "status": str(item.get("status") or ""),
                    "error": error_text,
                },
            }
        )
    db.append_learning_logs(rows)


def _external_source_cache_path(*, output_dir: Path, drama: dict, video_url: str, safe_title: str) -> Path:
    digest = hashlib.sha1(video_url.encode("utf-8")).hexdigest()[:10]
    app_id = str(drama.get("app_id") or drama.get("source_platform") or "external").strip() or "external"
    return output_dir / f"{safe_title}_{app_id}_{digest}_external_source.mp4"


def _ensure_external_source_file(*, video_url: str, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists() and output_path.stat().st_size > 0:
        return output_path
    return Path(download_segment_video(video_url, output_path=output_path)).expanduser().resolve()


def _is_hls_url(video_url: str) -> bool:
    normalized = str(video_url or "").strip().lower()
    return ".m3u8" in normalized or normalized.startswith("hls://")


def _download_remote_media_via_ffmpeg(*, video_url: str, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    command = [
        "ffmpeg",
        "-y",
        "-protocol_whitelist",
        "file,http,https,tcp,tls,crypto",
        "-i",
        str(video_url),
        "-map",
        "0:v:0",
        "-map",
        "0:a?",
        "-c:v",
        "libx264",
        "-preset",
        "veryfast",
        "-crf",
        "23",
        "-pix_fmt",
        "yuv420p",
        "-c:a",
        "aac",
        "-ar",
        "44100",
        "-ac",
        "2",
        "-movflags",
        "+faststart",
        str(output_path),
    ]
    try:
        subprocess.run(command, capture_output=True, text=True, check=True)
    except FileNotFoundError as exc:
        raise RuntimeError("系统未安装 ffmpeg，无法下载 HLS/远程官方素材") from exc
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(f"ffmpeg 下载远程素材失败: {(exc.stderr or '').strip()}") from exc
    if not output_path.exists() or output_path.stat().st_size <= 0:
        raise RuntimeError("ffmpeg 下载远程素材后文件为空")
    return output_path


def _ensure_remote_media_source(*, video_url: str, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists() and output_path.stat().st_size > 0:
        return output_path
    if _is_hls_url(video_url):
        return _download_remote_media_via_ffmpeg(video_url=video_url, output_path=output_path)
    return _ensure_external_source_file(video_url=video_url, output_path=output_path)


def _external_segment_window(*, drama: dict, clip_duration: int) -> tuple[int, int]:
    total_duration = 0
    for value in (
        drama.get("external_video_duration_seconds"),
        ((drama.get("raw") or {}).get("external_video_duration_seconds") if isinstance(drama.get("raw"), dict) else 0),
    ):
        try:
            total_duration = max(total_duration, int(value or 0))
        except (TypeError, ValueError):
            continue
    if total_duration <= 0:
        try:
            total_duration = int(probe_video(str(drama.get("cached_source_path") or "")).get("file_duration") or 0)
        except Exception:
            total_duration = 0

    target_duration = max(12, min(30, int(clip_duration or 15)))
    if total_duration > 0:
        target_duration = min(target_duration, total_duration)
    capacity = max(1, min(60, int(drama.get("external_estimated_output_count") or 1)))
    slot_index = max(0, int(drama.get("reuse_index") or 0))
    max_start = max(0, total_duration - target_duration)
    if max_start <= 0 or capacity <= 1:
        return 0, target_duration
    start = min(max_start, max(0, slot_index * target_duration))
    return max(0, start), target_duration


def _external_segment_output_path(
    *,
    output_dir: Path,
    safe_title: str,
    account: dict[str, object],
    start_seconds: int,
    duration_seconds: int,
) -> Path:
    account_id = str(account.get("account_id") or "account").strip() or "account"
    digest = hashlib.sha1(
        f"{safe_title}|{account_id}|{int(start_seconds)}|{int(duration_seconds)}|{time.time_ns()}".encode("utf-8")
    ).hexdigest()[:10]
    return output_dir / f"{safe_title}_{account_id}_{duration_seconds}s_{digest}_external_cut.mp4"


def _validate_video_decode(path: Path) -> tuple[bool, str]:
    command = [
        "ffmpeg",
        "-v",
        "error",
        "-xerror",
        "-err_detect",
        "explode",
        "-i",
        str(path),
        "-map",
        "0:v:0",
        "-map",
        "0:a?",
        "-f",
        "null",
        "-",
    ]
    try:
        completed = subprocess.run(command, capture_output=True, text=True, check=True)
    except FileNotFoundError as exc:
        raise RuntimeError("系统未安装 ffmpeg，无法校验实时榜外部素材片段") from exc
    except subprocess.CalledProcessError as exc:
        stderr = (exc.stderr or "").strip()
        return False, stderr or "ffmpeg 解码校验失败"
    stderr = (completed.stderr or "").strip()
    if stderr:
        return False, stderr
    return True, ""


def _run_external_segment_ffmpeg(
    *,
    source_path: Path,
    output_path: Path,
    start_seconds: int,
    duration_seconds: int,
    safe_mode: bool,
) -> None:
    if safe_mode:
        command = [
            "ffmpeg",
            "-y",
            "-i",
            str(source_path),
            "-vf",
            f"trim=start={max(0, int(start_seconds or 0))}:duration={max(1, int(duration_seconds or 1))},setpts=PTS-STARTPTS",
            "-af",
            f"atrim=start={max(0, int(start_seconds or 0))}:duration={max(1, int(duration_seconds or 1))},asetpts=PTS-STARTPTS",
            "-map",
            "0:v:0",
            "-map",
            "0:a?",
            "-c:v",
            "libx264",
            "-preset",
            "veryfast",
            "-crf",
            "23",
            "-pix_fmt",
            "yuv420p",
            "-c:a",
            "aac",
            "-ar",
            "44100",
            "-ac",
            "2",
            "-movflags",
            "+faststart",
            str(output_path),
        ]
    else:
        command = [
            "ffmpeg",
            "-y",
            "-i",
            str(source_path),
            "-ss",
            str(max(0, int(start_seconds or 0))),
            "-t",
            str(max(1, int(duration_seconds or 1))),
            "-map",
            "0:v:0",
            "-map",
            "0:a?",
            "-c:v",
            "libx264",
            "-preset",
            "veryfast",
            "-crf",
            "23",
            "-pix_fmt",
            "yuv420p",
            "-c:a",
            "aac",
            "-ar",
            "44100",
            "-ac",
            "2",
            "-movflags",
            "+faststart",
            str(output_path),
        ]
    subprocess.run(command, capture_output=True, text=True, check=True)


def _cut_external_video_segment(
    *,
    source_path: Path,
    output_path: Path,
    start_seconds: int,
    duration_seconds: int,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        _run_external_segment_ffmpeg(
            source_path=source_path,
            output_path=output_path,
            start_seconds=start_seconds,
            duration_seconds=duration_seconds,
            safe_mode=False,
        )
    except FileNotFoundError as exc:
        raise RuntimeError("系统未安装 ffmpeg，无法切分实时榜外部素材") from exc
    except subprocess.CalledProcessError as exc:
        stderr = (exc.stderr or "").strip()
        raise RuntimeError(f"ffmpeg 切分实时榜外部素材失败: {stderr}") from exc
    if not output_path.exists() or output_path.stat().st_size <= 0:
        raise RuntimeError("ffmpeg 切分后的实时榜片段为空")
    valid, validation_error = _validate_video_decode(output_path)
    if not valid:
        try:
            _run_external_segment_ffmpeg(
                source_path=source_path,
                output_path=output_path,
                start_seconds=start_seconds,
                duration_seconds=duration_seconds,
                safe_mode=True,
            )
        except subprocess.CalledProcessError as exc:
            stderr = (exc.stderr or "").strip()
            raise RuntimeError(
                f"ffmpeg 切分后的实时榜片段校验失败，回退重编码也失败: {validation_error or stderr}"
            ) from exc
        valid, validation_error = _validate_video_decode(output_path)
        if not valid:
            raise RuntimeError(f"ffmpeg 切分后的实时榜片段损坏: {validation_error}")
    return output_path


def _clip_batch_item(item: dict, args: argparse.Namespace, config) -> dict:
    drama = item["drama"]
    episode_order = int((item.get("episode") or {}).get("episode_order") or 1)
    clip_options = item["clip_options"]
    source_mode = str(drama.get("source_mode") or "official").strip() or "official"
    external_video_url = _external_video_url(drama)
    output_dir = Path(args.download_dir).expanduser().resolve() if args.download_dir else Path(config.clipped_dir) / "batch"
    output_dir.mkdir(parents=True, exist_ok=True)

    if source_mode in {"external_video", "official_ffmpeg"}:
        media_url = external_video_url if source_mode == "external_video" else str((item.get("episode") or {}).get("play_url") or "").strip()
        if not media_url:
            raise RuntimeError("外部/官方 ffmpeg 素材缺少可下载 media_url")
        safe_title = re.sub(r"[\\\\/:*?\"<>|]+", "_", str(drama.get("title") or drama.get("serial_id") or "external")).strip() or "external"
        raw_source_path = _external_source_cache_path(
            output_dir=output_dir,
            drama=drama,
            video_url=media_url,
            safe_title=safe_title,
        )
        downloaded_source = _ensure_remote_media_source(
            video_url=media_url,
            output_path=raw_source_path,
        )
        drama["cached_source_path"] = str(downloaded_source)
        clip_duration = clip_options.get("duration") or 90
        try:
            clip_duration_int = int(float(clip_duration))
        except (TypeError, ValueError):
            clip_duration_int = 90
        start_seconds, duration_seconds = _external_segment_window(
            drama=drama,
            clip_duration=clip_duration_int,
        )
        downloaded_file = _cut_external_video_segment(
            source_path=downloaded_source,
            output_path=_external_segment_output_path(
                output_dir=output_dir,
                safe_title=safe_title,
                account=item.get("account") or {},
                start_seconds=start_seconds,
                duration_seconds=duration_seconds,
            ),
            start_seconds=start_seconds,
            duration_seconds=duration_seconds,
        )
        publish_ready_file = _ensure_vertical_publish_file(
            str(downloaded_file),
            target_width=config.clip_target_width,
            target_height=config.clip_target_height,
        )
        clip_result = {
            "task": {
                "key": "external_video_ffmpeg_segment" if source_mode == "external_video" else "official_ffmpeg_segment",
                "params": {
                    "start_seconds": start_seconds,
                    "duration_seconds": duration_seconds,
                    "cut_type": "ffmpeg_segment",
                },
            },
            "submit": {
                "provider": "external_video_ffmpeg_segment" if source_mode == "external_video" else "official_ffmpeg_segment",
                "task_id": "",
                "manus_id": "",
                "response": {
                    "source_path": str(downloaded_source),
                    "start_seconds": start_seconds,
                    "duration_seconds": duration_seconds,
                },
            },
            "manus_id": "",
            "manus_status": "local_segment_ready",
            "downloaded_file": str(downloaded_file),
            "publish_ready_file": str(publish_ready_file),
            "downloaded_metadata": _probe_video_safely(str(downloaded_file)),
            "publish_ready_metadata": _probe_video_safely(str(publish_ready_file)),
            "execution_provider": "external_video_ffmpeg_segment" if source_mode == "external_video" else "official_ffmpeg_segment",
            "source_upload_id": "",
            "source_window_id": "",
            "source_clip_path": str(downloaded_source),
            "media_url": media_url,
            "segment_start_seconds": start_seconds,
            "segment_duration_seconds": duration_seconds,
        }
        return {
            **item,
            "status": "clipped",
            "source_path": str(downloaded_source),
            "clip": clip_result,
        }

    clip_args = SimpleNamespace(
        task_id=drama.get("task_id"),
        search="",
        serial_id=drama.get("serial_id"),
        app_id=drama.get("app_id"),
        episode_order=episode_order,
        drama_task_type=drama.get("task_type") or "1",
        deduplication=[],
        watermark="",
        duration="",
        output_count=1,
        cut_type="ai_cut_animation",
        script_count=1,
        merge_video=False,
        upload_timeout=args.upload_timeout,
        submit_timeout=args.submit_timeout,
        timeout=args.timeout,
        poll_interval=args.poll_interval,
        source_prepare_retry_count=getattr(args, "source_prepare_retry_count", 0),
    )
    source_context = resolve_drama_episode_context(clip_args)
    app_id = str(drama.get("app_id") or "").strip()
    serial_id = str(drama.get("serial_id") or "").strip()
    third_serial_id = str(
        drama.get("third_serial_id")
        or ((drama.get("raw") or {}).get("third_serial_id") if isinstance(drama.get("raw"), dict) else "")
    ).strip()
    if not app_id or not third_serial_id:
        raise RuntimeError("ai-cut 剪辑缺少 app_id 或 third_serial_id")
    task_create = create_short_drama_clip_task(
        app_id=app_id,
        third_serial_ids=[third_serial_id],
        auto_clip_output_folder="barry_video_batch",
        download_output_folder="barry_video_batch",
        source=str(clip_options.get("source") or DEFAULT_AI_ANIMATION_SOURCE),
        max_episodes_per_serial=max(
            int(clip_options.get("max_episodes_per_serial") or DEFAULT_MAX_EPISODES_PER_SERIAL),
            int(episode_order or 0),
        ),
        auto_clip_enabled=bool(
            clip_options.get("auto_clip_enabled")
            if clip_options.get("auto_clip_enabled") is not None
            else DEFAULT_AUTO_CLIP_ENABLED
        ),
        template_id=int(clip_options.get("template_id") or DEFAULT_TEMPLATE_ID),
        segment_seconds=int(clip_options.get("segment_seconds") or DEFAULT_SEGMENT_SECONDS),
        segment_max_seconds=int(clip_options.get("segment_max_seconds") or DEFAULT_SEGMENT_MAX_SECONDS),
        process_concurrency=int(clip_options.get("process_concurrency") or DEFAULT_PROCESS_CONCURRENCY),
        max_total_duration_seconds=int(
            clip_options.get("max_total_duration_seconds") or DEFAULT_MAX_TOTAL_DURATION_SECONDS
        ),
        use_auto_migration=bool(
            clip_options.get("use_auto_migration")
            if clip_options.get("use_auto_migration") is not None
            else DEFAULT_USE_AUTO_MIGRATION
        ),
        auto_clip_sync_material_agent_id=int(
            clip_options.get("auto_clip_sync_material_agent_id") or DEFAULT_SYNC_MATERIAL_AGENT_ID
        ),
        auto_clip_sync_material_category_id=int(
            clip_options.get("auto_clip_sync_material_category_id") or DEFAULT_SYNC_MATERIAL_CATEGORY_ID
        ),
        force_download=True,
        timeout=args.submit_timeout,
    )
    task_id = str(task_create.get("task_id") or "").strip()
    if not task_id:
        raise RuntimeError(f"ai-cut 创建任务未返回 task_id: {json.dumps(task_create, ensure_ascii=False)}")
    task_body = wait_for_short_drama_clip_task(
        task_id,
        timeout=args.timeout,
        poll_interval=args.poll_interval,
        request_timeout=args.submit_timeout,
    )
    serial_payload = None
    for payload in (task_body.get("serials") or []):
        if str((payload or {}).get("third_serial_id") or "").strip() == third_serial_id:
            serial_payload = dict(payload)
            break
    task_body, refreshed_serial_payload, chosen_segment = wait_for_serial_success_segment(
        task_id,
        third_serial_id=third_serial_id,
        preferred_episode_order=episode_order,
        timeout=args.timeout,
        poll_interval=args.poll_interval,
        request_timeout=args.submit_timeout,
        initial_task_body=task_body,
    )
    serial_payload = refreshed_serial_payload or serial_payload
    if not serial_payload:
        raise RuntimeError(f"ai-cut 未返回 serial 结果: {third_serial_id}")
    if not chosen_segment:
        raise RuntimeError(describe_serial_failure(serial_payload))

    safe_title = re.sub(r"[\\\\/:*?\"<>|]+", "_", str(drama.get("title") or serial_id)).strip() or serial_id
    local_clip_path = download_segment_video(
        str(chosen_segment.get("video_url") or ""),
        output_path=output_dir / f"{safe_title}_E{episode_order:02d}_aicut.mp4",
    )
    publish_ready_file = _ensure_vertical_publish_file(
        local_clip_path,
        target_width=config.clip_target_width,
        target_height=config.clip_target_height,
    )
    clip_result = {
        "task": {
            "key": "ai_cut_animation",
            "params": {
                "app_id": app_id,
                "third_serial_ids": [third_serial_id],
                "preferred_episode_order": episode_order,
                "template_id": int(clip_options.get("template_id") or DEFAULT_TEMPLATE_ID),
                "segment_seconds": int(clip_options.get("segment_seconds") or DEFAULT_SEGMENT_SECONDS),
                "segment_max_seconds": int(clip_options.get("segment_max_seconds") or DEFAULT_SEGMENT_MAX_SECONDS),
                "process_concurrency": int(
                    clip_options.get("process_concurrency") or DEFAULT_PROCESS_CONCURRENCY
                ),
                "max_total_duration_seconds": int(
                    clip_options.get("max_total_duration_seconds") or DEFAULT_MAX_TOTAL_DURATION_SECONDS
                ),
                "max_episodes_per_serial": max(
                    int(clip_options.get("max_episodes_per_serial") or DEFAULT_MAX_EPISODES_PER_SERIAL),
                    int(episode_order or 0),
                ),
                "source": str(clip_options.get("source") or DEFAULT_AI_ANIMATION_SOURCE),
                "auto_clip_enabled": bool(
                    clip_options.get("auto_clip_enabled")
                    if clip_options.get("auto_clip_enabled") is not None
                    else DEFAULT_AUTO_CLIP_ENABLED
                ),
                "use_auto_migration": bool(
                    clip_options.get("use_auto_migration")
                    if clip_options.get("use_auto_migration") is not None
                    else DEFAULT_USE_AUTO_MIGRATION
                ),
                "auto_clip_sync_material_agent_id": int(
                    clip_options.get("auto_clip_sync_material_agent_id") or DEFAULT_SYNC_MATERIAL_AGENT_ID
                ),
                "auto_clip_sync_material_category_id": int(
                    clip_options.get("auto_clip_sync_material_category_id") or DEFAULT_SYNC_MATERIAL_CATEGORY_ID
                ),
            },
        },
        "submit": {
            "provider": "ai_cut_animation",
            "task_id": task_id,
            "manus_id": task_id,
            "response": task_create,
        },
        "manus_id": task_id,
        "manus_status": task_body.get("status"),
        "downloaded_file": str(local_clip_path),
        "publish_ready_file": str(publish_ready_file),
        "downloaded_metadata": _probe_video_safely(str(local_clip_path)),
        "publish_ready_metadata": _probe_video_safely(str(publish_ready_file)),
        "execution_provider": "ai_cut_animation",
        "ai_animation_task_id": task_id,
        "ai_animation_download_status": serial_payload.get("download_status"),
        "ai_animation_clip_status": ((serial_payload.get("clip") or {}).get("status") if isinstance(serial_payload.get("clip"), dict) else ""),
        "ai_animation_segment": chosen_segment,
    }
    return {
        **item,
        "status": "clipped",
        "source": source_context,
        "clip": clip_result,
    }


def _promotion_caption(item: dict, platform: str) -> dict[str, str]:
    existing = item.get("promotion") if isinstance(item.get("promotion"), dict) else {}
    if existing:
        return {
            "caption": str(existing.get("caption") or item.get("caption") or "").strip(),
            "promotion_link": str(existing.get("promotion_link") or "").strip(),
            "promotion_code": str(existing.get("promotion_code") or "").strip(),
            "promote_code_content": str(existing.get("promote_code_content") or "").strip(),
            "promotion_platform_id": str(existing.get("promotion_platform_id") or "").strip(),
        }

    drama = item["drama"]
    promotion_platform = 1 if _line_in("yourchannel") else PUBLISH_TO_PROMOTION_PLATFORM.get(platform)
    if not promotion_platform:
        return {"caption": str(drama.get("title") or ""), "promotion_link": "", "promotion_code": ""}

    def _fetch_link_payload(task_id: str, task_type: str, label: str) -> dict:
        payload = require_success(
            receive_task(
                task_id=task_id,
                task_type=task_type,
                platform=promotion_platform,
            ),
            label,
        )
        atr_id = payload.get("atr_id")
        if atr_id:
            require_success(
                active_task(atr_id),
                f"激活 {PROMOTION_PLATFORMS[promotion_platform]} 推广任务",
            )
        return payload

    payload = None
    primary_task_id = str(drama.get("task_id") or "").strip()
    matched_official_task_id = str(drama.get("matched_official_task_id") or "").strip()
    if primary_task_id:
        try:
            payload = _fetch_link_payload(
                primary_task_id,
                str(drama.get("task_type") or "1"),
                f"获取 {PROMOTION_PLATFORMS[promotion_platform]} 推广链接",
            )
        except Exception:
            payload = None
    if payload is None:
        anchor = drama.get("promotion_anchor") if isinstance(drama.get("promotion_anchor"), dict) else {}
        anchor_task_id = str(anchor.get("task_id") or "").strip()
        if not anchor_task_id:
            raise RuntimeError("缺少可用推广任务 task_id，无法获取推广链接")
        payload = _fetch_link_payload(
            anchor_task_id,
            str(anchor.get("task_type") or "1"),
            f"获取实时榜锚点 {PROMOTION_PLATFORMS[promotion_platform]} 推广链接",
        )

    link_entry = build_promotion_link_entry(promotion_platform, payload)
    promotion_link = (
        link_entry.get("serial_link")
        or link_entry.get("app_link")
        or link_entry.get("tiktok_url")
        or ""
    )
    matched_official_used = bool(matched_official_task_id and primary_task_id == matched_official_task_id)
    use_realtime_app_fallback = _candidate_fetch_source(drama) == "realtime_rank_external" and not matched_official_used
    if not use_realtime_app_fallback:
        caption = str(link_entry.get("promote_code_content") or "").strip() or promotion_link or str(drama.get("title") or "")
    else:
        app_label = _app_label(str(drama.get("app_id") or drama.get("source_platform") or "")).strip() or str(drama.get("app_id") or "").strip() or "series"
        drama_title = str(drama.get("title") or "").strip() or "this series"
        caption = "\n".join(
            line
            for line in (
                f"watch👉🏻 {promotion_link}".strip() if promotion_link else "",
                "🌟 Continue the story here",
                f'👉🏻 📲 Find the full series on the "{app_label}"app',
                f'🔍 Look up "{drama_title}", to enjoy every episode✨!',
            )
            if line
        ).strip() or promotion_link or drama_title
    return {
        "caption": caption,
        "promotion_link": promotion_link,
        "promotion_code": str(link_entry.get("code") or ""),
        "promote_code_content": str(link_entry.get("promote_code_content") or "").strip(),
        "promotion_platform_id": str(promotion_platform or ""),
    }


def _publish_batch_item(item: dict, args: argparse.Namespace, platform: str, attempt: int) -> dict:
    clip = item.get("clip") or {}
    account = item.get("account") or {}
    publish_ready_file = str(clip.get("publish_ready_file") or "")
    if not publish_ready_file:
        exc = RuntimeError("缺少待发布成片")
        attempts = list(item.get("publish_attempts") or [])
        attempts.append({"attempt": attempt, "status": "failed", "error": str(exc)})
        return {**item, "status": "failed", "error": str(exc), "publish_attempts": attempts}
    try:
        drama = item.get("drama") if isinstance(item.get("drama"), dict) else {}
        validate_source_episode_constraints(
            item.get("episode") if isinstance(item.get("episode"), dict) else {},
            source_mode=str(drama.get("source_mode") or ""),
        )
        refreshed_meta = validate_publish_clip_constraints(clip)
        if refreshed_meta:
            clip = {**clip, **refreshed_meta}
        promotion = _promotion_caption(item, platform)
        validate_promotion_constraints(platform, promotion)
        upload_context = upload_publish_file(publish_ready_file)
        payload = {
            "team_id": str(account.get("team_id") or ""),
            "text": promotion["caption"],
            "file_url": str(upload_context.get("publish_file_url") or ""),
            "post_status": 0,
            "social_type": platform,
        }
        if platform in {"FACEBOOK", "INSTAGRAM"}:
            payload["type"] = "REEL"
        body = require_success(create_publish_post(payload), "发布帖子")
        tasks = body.get("tasks") if isinstance(body.get("tasks"), list) else []
        if not tasks:
            tasks = [
                {
                    "team_id": payload["team_id"],
                    "task_id": body.get("task_id"),
                    "post_id": body.get("post_id"),
                    "status": body.get("status") or "SUBMITTED",
                }
            ]
        attempts = list(item.get("publish_attempts") or [])
        attempts.append(
            {
                "attempt": attempt,
                "status": "submitted",
                "task_count": len(tasks),
                "team_id": payload["team_id"],
            }
        )
        return {
            **item,
            "status": "published_submitted",
            "error": "",
            "clip": clip,
            "promotion": promotion,
            "publish_attempts": attempts,
            "publish_retry_count": max(0, len(attempts) - 1),
            "publish": {
                "payload": payload,
                "upload": upload_context,
                "response": body,
                "tasks": tasks,
            },
        }
    except Exception as exc:
        attempts = list(item.get("publish_attempts") or [])
        attempts.append({"attempt": attempt, "status": "failed", "error": str(exc)})
        return {
            **item,
            "status": "failed",
            "error": str(exc),
            "publish_attempts": attempts,
            "publish_retry_count": max(0, len(attempts) - 1),
        }


def _publish_batch_with_retries(
    items: list[dict],
    args: argparse.Namespace,
    platform: str,
    *,
    max_attempts: int | None = None,
) -> tuple[list[dict], list[dict]]:
    state_by_index = {int(item.get("index") or 0): dict(item) for item in items}
    latest_records: list[dict] = []
    max_attempts = max_attempts if max_attempts is not None else max(1, int(args.publish_retries) + 1)
    max_attempts = max(1, int(max_attempts))
    pending = [dict(item) for item in items]

    for attempt in range(1, max_attempts + 1):
        if not pending:
            break
        attempted = _run_parallel(
            pending,
            max(1, int(args.publish_concurrency)),
            lambda item: _publish_batch_item(item, args, platform, attempt),
        )
        for item in attempted:
            _merge_publish_attempt_result(state_by_index, item)

        all_tasks = [
            task
            for item in state_by_index.values()
            for task in (((item.get("publish") or {}).get("tasks")) or [])
        ]
        latest_records = _poll_local_publish_records(
            platform=platform,
            tasks=all_tasks,
            wait_seconds=args.collect_wait_seconds,
            poll_interval=args.collect_poll_interval,
        ) if all_tasks else []
        if attempt >= max_attempts:
            break
        records_by_key = _record_by_task_key(latest_records)
        pending = [
            dict(item)
            for _, item in sorted(state_by_index.items())
            if _item_should_retry_publish(item, records_by_key)
        ]
        pending.sort(key=lambda item: int(item.get("index") or 0))

    return [item for _, item in sorted(state_by_index.items())], latest_records


def _batch_report_zh(payload: dict) -> dict:
    is_dry_run = str(payload.get("status") or "").lower() == "dry_run"
    items = payload.get("items") if isinstance(payload.get("items"), list) else []
    records = payload.get("publish_records") if isinstance(payload.get("publish_records"), list) else []
    records_by_key = _record_by_task_key(records)
    cleanup = payload.get("cleanup") if isinstance(payload.get("cleanup"), dict) else {}
    cleanup_deleted = {str(path) for path in cleanup.get("deleted_paths", [])}
    item_reports = [_publish_item_report(item, records_by_key, cleanup_deleted) for item in items]
    if is_dry_run:
        for report in item_reports:
            report["发布情况"] = "待执行"
            report["是否可自动重试"] = ""
            report["处理建议"] = ""
            report["失败原因"] = ""
            report["错误"] = ""
    success_reports = [
        report
        for report in item_reports
        if _is_success_publish_outcome(report)
    ]
    processing_reports = [report for report in item_reports if _is_processing_publish_outcome(report)]
    failed_reports = [] if is_dry_run else [report for report in item_reports if _is_failed_publish_outcome(report)]
    theater_counts: dict[str, int] = {}
    for report in item_reports:
        theater = str(report.get("剧场") or "").strip()
        if theater:
            theater_counts[theater] = theater_counts.get(theater, 0) + 1
    safety_gate = payload.get("safety_gate") if isinstance(payload.get("safety_gate"), dict) else {}
    strategy_memory = payload.get("strategy_memory") if isinstance(payload.get("strategy_memory"), dict) else {}
    return {
        "执行模式": "批量短剧剪辑发布（仅规划）" if is_dry_run else "批量短剧剪辑发布",
        "目标平台": _platform_label(str(payload.get("platform") or "")),
        "请求数量": payload.get("requested_count"),
        "计划数量": len(items),
        "计划缺口数": int(payload.get("planned_shortfall_count") or 0),
        "可用唯一短剧数": int(payload.get("unique_playable_source_count") or 0),
        "实时榜外部素材数": int(payload.get("realtime_external_unique_count") or 0),
        "实时榜外部素材填充槽位数": int(payload.get("realtime_external_slot_fill_count") or 0),
        "复用补量数": int(payload.get("source_reuse_fill_count") or 0),
        "剧场分布": theater_counts,
        "剪辑成功数": len([item for item in items if item.get("clip")]),
        "发布提交数": sum(len(((item.get("publish") or {}).get("tasks") or [])) for item in items),
        "发布成功数": len(success_reports),
        "发布处理中数": 0 if is_dry_run else len(processing_reports),
        "失败数": len(failed_reports),
        "阶段耗时": payload.get("timing_zh") or {},
        "安全门槛": {
            "通过数": int(safety_gate.get("passed_count") or len(items)),
            "拦截数": int(safety_gate.get("rejected_count") or 0),
            "补位成功数": int(safety_gate.get("replacement_filled_count") or 0),
            "备用候选数": int(safety_gate.get("reserve_source_count") or 0),
            "补位尝试数": int(safety_gate.get("reserve_attempt_count") or 0),
            "补位缺口数": int(safety_gate.get("unfilled_count") or 0),
            "拦截预览": list(safety_gate.get("rejected_preview") or []),
            "拦截明细": list(safety_gate.get("rejected_details") or []),
        },
        "策略记忆": strategy_memory,
        "发布成功视频": success_reports,
        "账号发布结果": [
            {
                "账号": report.get("账号"),
                "平台": report.get("平台"),
                "短剧": report.get("短剧"),
                "集数": report.get("集数"),
                "发布情况": report.get("发布情况"),
                "剪辑手法": report.get("剪辑手法"),
                "去重手法": report.get("去重手法"),
                "失败原因": report.get("失败原因"),
            }
            for report in item_reports
        ],
        "发布失败任务": failed_reports,
        "任务明细": item_reports,
    }


def _batch_detail_block_zh(item: dict) -> str:
    result = str(item.get("发布情况") or "")
    index = str(item.get("序号") or "").strip()
    short_drama = str(item.get("短剧") or "").strip()
    episode = int(item.get("集数") or 0)
    local_source = str(item.get("本地源视频") or "").strip()
    lines = [
        _join_non_empty(
            [
                f"{index}." if index else "",
                str(item.get("账号") or ""),
                str(item.get("平台") or ""),
                result,
            ],
            sep=" ",
        ),
        _join_non_empty(
            [
                f"剪辑：{item.get('剪辑手法')}",
                f"去重：{item.get('去重手法')}",
                f"比例：{item.get('目标比例')}",
            ],
            sep=" | ",
        ),
        _join_non_empty(
            [
                f"视频：{item.get('视频时长')}",
                str(item.get("视频分辨率") or ""),
                str(item.get("文件大小") or ""),
            ],
            sep=" | ",
        ),
    ]
    source_line = ""
    if local_source:
        source_line = f"来源视频：{local_source}"
    elif short_drama:
        source_line = _join_non_empty(
            [
                f"剧集：《{short_drama}》" + (f"第 {episode} 集" if episode > 0 else ""),
                str(item.get("剧场") or ""),
                str(item.get("语言") or ""),
            ],
            sep=" | ",
        )
    if source_line:
        lines.insert(1, source_line)
    publish_time = str(item.get("发布时间") or "").strip()
    if publish_time:
        lines.append(f"发布：{publish_time}")
    metrics = _join_non_empty(
        [
            f"播放 {item.get('播放量')}",
            f"点赞 {item.get('点赞数')}",
            f"评论 {item.get('评论数')}",
            f"分享 {item.get('分享数')}",
        ],
        sep="，",
    )
    if metrics:
        lines.append(f"数据：{metrics}")
    post_id = str(item.get("平台帖子ID") or "").strip()
    if post_id:
        lines.append(f"帖子ID：{post_id}")
    local_status = str(item.get("本地成片状态") or "").strip()
    if local_status:
        lines.append(f"本地成片：{local_status}")
    if result != "发布成功":
        failure_reason = str(item.get("失败原因") or item.get("错误") or "").strip()
        if failure_reason:
            lines.append(f"失败原因：{failure_reason}")
        advice = str(item.get("处理建议") or "").strip()
        if advice:
            lines.append(f"建议：{advice}")
    return "\n   ".join(line.rstrip("。；，, ") for line in lines if str(line or "").strip())


def _batch_user_summary_zh(report: dict) -> str:
    mode = str(report.get("执行模式") or "")
    if "仅规划" in mode:
        lines: list[str] = [
            _join_non_empty(
                [
                    f"本次只完成了批量任务规划，目标平台 {report.get('目标平台')}",
                    f"请求 {report.get('请求数量')} 条",
                    f"实际计划 {report.get('计划数量')} 条",
                ]
            )
            + "，尚未执行剪辑和发布。",
        ]
        theater_counts = report.get("剧场分布") if isinstance(report.get("剧场分布"), dict) else {}
        if theater_counts:
            theater_line = "，".join(f"{theater} {count} 条" for theater, count in theater_counts.items())
            if theater_line:
                lines.append(f"本轮剧场分布：{theater_line}。")
        external_slots = int(report.get("实时榜外部素材填充槽位数") or 0)
        external_unique = int(report.get("实时榜外部素材数") or 0)
        if external_slots or external_unique:
            lines.append(f"实时榜外部素材：命中 {external_unique} 个素材，优先填充 {external_slots} 个账号槽位。")
        timing = report.get("阶段耗时") if isinstance(report.get("阶段耗时"), dict) else {}
        if timing:
            timing_line = "，".join(f"{key} {value}" for key, value in timing.items())
            if timing_line:
                lines.append(f"规划阶段耗时：{timing_line}。")
        safety_gate = report.get("安全门槛") if isinstance(report.get("安全门槛"), dict) else {}
        if safety_gate:
            lines.append(
                f"安全门槛：通过 {int(safety_gate.get('通过数') or 0)} 条，拦截 {int(safety_gate.get('拦截数') or 0)} 条。"
            )
        lines.append("如需真实执行剪辑与发布，请使用 --execute。")
        return "\n".join(line for line in lines if str(line or "").strip())

    lines: list[str] = [_report_conclusion_lines(report)[0]]
    theater_counts = report.get("剧场分布") if isinstance(report.get("剧场分布"), dict) else {}
    if theater_counts:
        theater_line = "，".join(
            f"{theater} {count} 条"
            for theater, count in theater_counts.items()
        )
        if theater_line:
            lines.append(f"本轮剧场分布：{theater_line}。")
    external_slots = int(report.get("实时榜外部素材填充槽位数") or 0)
    external_unique = int(report.get("实时榜外部素材数") or 0)
    if external_slots or external_unique:
        lines.append(f"实时榜外部素材：命中 {external_unique} 个素材，优先填充 {external_slots} 个账号槽位。")
    timing = report.get("阶段耗时") if isinstance(report.get("阶段耗时"), dict) else {}
    if timing:
        timing_line = "，".join(f"{key} {value}" for key, value in timing.items())
        if timing_line:
            lines.append(f"阶段耗时：{timing_line}。")
    safety_gate = report.get("安全门槛") if isinstance(report.get("安全门槛"), dict) else {}
    if safety_gate:
        lines.append(
            f"安全门槛：通过 {int(safety_gate.get('通过数') or 0)} 条，拦截 {int(safety_gate.get('拦截数') or 0)} 条。"
        )
    strategy_memory = report.get("策略记忆") if isinstance(report.get("策略记忆"), dict) else {}
    cooldown_days = int(strategy_memory.get("cooldown_days") or 0)
    cooldown_serial_count = int(strategy_memory.get("cooldown_serial_count") or 0)
    if cooldown_days > 0:
        lines.append(f"策略记忆：最近 {cooldown_days} 天冷却中的短剧 {cooldown_serial_count} 部。")

    failed_reports = report.get("发布失败任务") if isinstance(report.get("发布失败任务"), list) else []
    retryable_failed = [
        item
        for item in failed_reports
        if str(item.get("是否可自动重试") or "") != "否"
    ]
    non_retryable_failed = [
        item
        for item in failed_reports
        if str(item.get("是否可自动重试") or "") == "否"
    ]
    if failed_reports:
        lines.append(
            f"未成功任务共 {len(failed_reports)} 条，其中可继续重试 {len(retryable_failed)} 条，不建议自动重试 {len(non_retryable_failed)} 条。"
        )

    retry_prompt = _failed_publish_prompt_zh(report)
    if retry_prompt:
        lines.append(retry_prompt)

    return "\n".join(lines)


def cmd_run_batch_drama(args) -> None:
    run_started_at = time.perf_counter()
    timings: dict[str, float] = {}
    config = load_config(args.config)
    ensure_runtime_dirs(config)
    if getattr(args, "source_prepare_retry_count", None) is None:
        args.source_prepare_retry_count = config.source_prepare_retry_count
    if args.clip_concurrency is None:
        args.clip_concurrency = config.clip_execute_concurrency
    if args.publish_concurrency is None:
        args.publish_concurrency = config.publish_execute_concurrency
    plan_heartbeat = _start_stage_heartbeat("选剧与剧集预检", detail=f"目标 {args.count} 条，发布到 {args.publish_platform}")
    try:
        plan_started_at = time.perf_counter()
        plan = _build_batch_plan(args, config)
        timings["选剧与剧集预检"] = time.perf_counter() - plan_started_at
    finally:
        _stop_stage_heartbeat("选剧与剧集预检", *plan_heartbeat)
    items = list(plan["items"])

    safety_heartbeat = _start_stage_heartbeat("安全门槛", detail=f"{len(items)} 条候选待校验")
    try:
        safety_started_at = time.perf_counter()
        approved_items, rejected_items = _run_batch_safety_gate(
            items,
            str(plan["platform"]),
            prefetch_promotion=not args.dry_run,
        )
        approved_items, rejected_items, safety_refill = _backfill_batch_safety_gate(
            approved_items=approved_items,
            rejected_items=rejected_items,
            reserve_sources=list(plan.get("reserve_sources") or []),
            args=args,
            platform=str(plan["platform"]),
            prefetch_promotion=not args.dry_run,
        )
        timings["安全门槛"] = time.perf_counter() - safety_started_at
    finally:
        _stop_stage_heartbeat("安全门槛", *safety_heartbeat)
    safety_gate = {
        "passed_count": len(approved_items),
        "rejected_count": len(rejected_items),
        "rejected_preview": list((safety_refill.get("rejected_details") or [])[:10]),
        "rejected_details": list(safety_refill.get("rejected_details") or []),
        "replacement_filled_count": int(safety_refill.get("replacement_filled_count") or 0),
        "reserve_source_count": int(safety_refill.get("reserve_source_count") or 0),
        "reserve_attempt_count": int(safety_refill.get("reserve_attempt_count") or 0),
        "unfilled_count": int(safety_refill.get("unfilled_count") or 0),
        "unfilled_slots": list(safety_refill.get("unfilled_slots") or []),
    }
    items = approved_items
    if not items:
        raise SystemExit(
            json.dumps(
                {
                    "status": "no_safe_batch_items",
                    "message": "安全门槛拦截后没有可继续剪辑发布的短剧。",
                    "safety_gate": safety_gate,
                },
                ensure_ascii=False,
                indent=2,
            )
        )

    if args.dry_run:
        payload = {
            "status": "dry_run",
            "mode": "batch_drama",
            "platform": plan["platform"],
            "requested_count": args.count,
            "dry_run_notice_zh": "当前仅完成规划，未执行剪辑和发布。如需真实执行，请加 --execute。",
            "items": items,
            "drama_platform_plan": plan.get("drama_platform_plan", []),
            "episode_precheck": plan.get("episode_precheck", {}),
            "safety_gate": safety_gate,
            "strategy_memory": plan.get("strategy_memory", {}),
            "skipped_preview": plan.get("skipped_preview", []),
            "unique_playable_source_count": int(plan.get("unique_playable_source_count") or 0),
            "source_reuse_fill_count": int(plan.get("source_reuse_fill_count") or 0),
            "realtime_external_unique_count": int(plan.get("realtime_external_unique_count") or 0),
            "realtime_external_slot_fill_count": int(plan.get("realtime_external_slot_fill_count") or 0),
            "planned_shortfall_count": max(0, int(plan.get("planned_shortfall_count") or 0) + max(0, int(args.count) - len(items))),
            "timings": timings,
            "timing_zh": _format_timing_zh(timings),
        }
        payload["report_zh"] = _batch_report_zh(payload)
        payload["user_summary_zh"] = _batch_user_summary_zh(payload["report_zh"])
        payload = _finalize_payload(payload)
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return

    clip_heartbeat = _start_stage_heartbeat(
        "剪辑与下载",
        detail=f"{len(items)} 条任务，剪辑并发 {max(1, int(args.clip_concurrency))}",
    )
    try:
        clip_started_at = time.perf_counter()
        clipped_items = _run_parallel(
            items,
            max(1, int(args.clip_concurrency)),
            lambda item: _clip_batch_item(item, args, config),
        )
        timings["剪辑与下载"] = time.perf_counter() - clip_started_at
    finally:
        _stop_stage_heartbeat("剪辑与下载", *clip_heartbeat)
    publishable_items = [item for item in clipped_items if item.get("status") != "failed" and item.get("clip")]
    publish_heartbeat = _start_stage_heartbeat(
        "上传发布与状态确认",
        detail=f"{len(publishable_items)} 条待发布，发布并发 {max(1, int(args.publish_concurrency))}",
    )
    try:
        publish_started_at = time.perf_counter()
        published_items, records = _publish_batch_with_retries(
            publishable_items,
            args,
            str(plan["platform"]),
        )
        timings["上传发布与状态确认"] = time.perf_counter() - publish_started_at
    finally:
        _stop_stage_heartbeat("上传发布与状态确认", *publish_heartbeat)
    failed_clip_items = [item for item in clipped_items if item.get("status") == "failed"]
    all_items = [*published_items, *failed_clip_items]
    all_items.sort(key=lambda item: int(item.get("index") or 0))

    cleanup_heartbeat = _start_stage_heartbeat("本地清理", detail="处理发布成功后的本地成片")
    try:
        cleanup_started_at = time.perf_counter()
        cleanup = {"deleted_paths": [], "errors": []}
        if not args.keep_output and records:
            successful_keys = {
                (str(record.get("team_id") or ""), str(record.get("task_id") or ""))
                for record in records
                if str(record.get("status") or "").upper() in SUCCESSFUL_PUBLISH_STATUSES
            }
            cleanup_paths: list[str] = []
            for item in all_items:
                tasks = ((item.get("publish") or {}).get("tasks")) or []
                if not tasks:
                    continue
                if all((str(task.get("team_id") or ""), str(task.get("task_id") or "")) in successful_keys for task in tasks):
                    clip = item.get("clip") or {}
                    cleanup_paths.extend([str(clip.get("downloaded_file") or ""), str(clip.get("publish_ready_file") or "")])
            cleanup = _cleanup_generated_files(cleanup_paths)
        timings["本地清理"] = time.perf_counter() - cleanup_started_at
    finally:
        _stop_stage_heartbeat("本地清理", *cleanup_heartbeat)
    timings["总耗时"] = time.perf_counter() - run_started_at

    payload = {
        "status": "done",
        "mode": "batch_drama",
        "platform": plan["platform"],
        "requested_count": args.count,
        "items": all_items,
        "publish_records": records,
        "drama_platform_plan": plan.get("drama_platform_plan", []),
        "episode_precheck": plan.get("episode_precheck", {}),
        "safety_gate": safety_gate,
        "strategy_memory": plan.get("strategy_memory", {}),
        "skipped_preview": plan.get("skipped_preview", []),
        "unique_playable_source_count": int(plan.get("unique_playable_source_count") or 0),
        "source_reuse_fill_count": int(plan.get("source_reuse_fill_count") or 0),
        "realtime_external_unique_count": int(plan.get("realtime_external_unique_count") or 0),
        "realtime_external_slot_fill_count": int(plan.get("realtime_external_slot_fill_count") or 0),
        "planned_shortfall_count": max(0, int(plan.get("planned_shortfall_count") or 0) + max(0, int(args.count) - len(items))),
        "cleanup": {
            "enabled": not args.keep_output,
            **cleanup,
        },
        "timings": timings,
        "timing_zh": _format_timing_zh(timings),
    }
    payload = _settle_publish_report_payload(
        payload,
        platform=str(plan["platform"]),
        wait_seconds=int(args.collect_wait_seconds),
        poll_interval=int(args.collect_poll_interval),
        settle_timeout_seconds=int(config.collect_settle_timeout_seconds),
        report_builder=_batch_report_zh,
    )
    try:
        _record_batch_learning_logs(
            round_id=None,
            items=all_items,
            safety_rejected=rejected_items,
            report_zh=payload["report_zh"],
            config=config,
        )
    except Exception as exc:
        payload["learning_log_warning"] = str(exc)
    payload["user_summary_zh"] = _batch_user_summary_zh(payload["report_zh"])
    payload["retry_prompt_zh"] = _failed_publish_prompt_zh(payload["report_zh"])
    failed_state = _failed_publish_state_payload(
        mode="batch_drama",
        platform=str(plan["platform"]),
        items=all_items,
        records=payload.get("publish_records") if isinstance(payload.get("publish_records"), list) else records,
    )
    _set_failed_publish_state(failed_state)
    payload = _finalize_payload(payload)
    print(json.dumps(payload, ensure_ascii=False, indent=2))
